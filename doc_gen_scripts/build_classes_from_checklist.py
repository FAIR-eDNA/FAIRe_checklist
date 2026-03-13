import os
from collections import defaultdict

from openpyxl import load_workbook
import yaml
from checklist_source import resolve_single_checklist_file

# Builds classes from the FAIRe checklist
SLOTS_DIR = "slots"
OUTPUT_PATH = "classes.yaml"
GLOSSARY_FILENAME = "glossary_annotation.yaml"


CLASS_DESCRIPTIONS = {
    "MetadataChecklist": "All FAIRe metadata fields.",
    "projectMetadata": "FAIRe Project-level metadata fields.",
    "sampleMetadata": "FAIRe Sample-level metadata including collection and environmental context.",
    "ampData": "FAIRe amplification-level assay output data fields.",
    "stdData": "FAIRe standard curve data used for quantification and calibration fields.",
    "experimentRunMetadata": "FAIRe sequencing run and library-level metadata fields.",
    "eLowQuantData": "FAIRe detection summaries for low-quantification outcome fields.",
    "taxaRaw": "FAIRe raw taxonomic assignment records before final curation fields.",
    "taxaFinal": "FAIRe final curated taxonomic assignment record fields.",
}


def split_pipe(raw):
    if not raw:
        return []
    return [x.strip() for x in str(raw).split("|") if x.strip()]


def load_checklist_rows_from_xlsx(path, sheet_name="checklist"):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Checklist Excel file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    header_names = [str(h).strip() if h is not None else "" for h in headers]
    out = []
    for row in rows:
        rec = {}
        for idx, key in enumerate(header_names):
            rec[key] = row[idx] if idx < len(row) else None
        out.append(rec)
    return out


def existing_slots():
    slots = {}
    for file_name in os.listdir(SLOTS_DIR):
        if not file_name.endswith(".yaml") or file_name == GLOSSARY_FILENAME:
            continue
        path = os.path.join(SLOTS_DIR, file_name)
        with open(path, "r", encoding="utf-8") as handle:
            doc = yaml.safe_load(handle) or {}
        if isinstance(doc, dict) and "name" in doc:
            slots[str(doc["name"]).strip()] = doc
    return slots


def preferred_class_order(class_names):
    preferred = [
        "projectMetadata",
        "sampleMetadata",
        "ampData",
        "stdData",
        "experimentRunMetadata",
        "eLowQuantData",
        "taxaRaw",
        "taxaFinal",
    ]
    ordered = [name for name in preferred if name in class_names]
    remainder = sorted([name for name in class_names if name not in preferred])
    return ordered + remainder


def main():
    checklist_xlsx = resolve_single_checklist_file(".xlsx")
    rows = load_checklist_rows_from_xlsx(checklist_xlsx)
    slots = existing_slots()
    slot_names = set(slots.keys())
    class_slots = defaultdict(set)

    # Primary source: checklist spreadsheet.
    for row in rows:
        slot_name = (row.get("term_name") or "").strip()
        if not slot_name or slot_name not in slot_names:
            continue
        for class_name in split_pipe(row.get("data_type")):
            class_slots[class_name].add(slot_name)

    # Fallback: slot annotations in case the worksheet lacks required columns.
    for slot_name, slot_def in slots.items():
        annotations = slot_def.get("annotations", {})
        if not isinstance(annotations, dict):
            continue
        data_types = annotations.get("data_type", [])
        if isinstance(data_types, str):
            data_types = split_pipe(data_types)
        if isinstance(data_types, list):
            for class_name in data_types:
                class_name = str(class_name).strip()
                if class_name:
                    class_slots[class_name].add(slot_name)

    classes_doc = {"classes": {}}

    all_slots_sorted = sorted(slot_names)
    classes_doc["classes"]["MetadataChecklist"] = {
        "description": CLASS_DESCRIPTIONS["MetadataChecklist"],
        "slots": all_slots_sorted,
    }

    for class_name in preferred_class_order(class_slots.keys()):
        classes_doc["classes"][class_name] = {
            "description": CLASS_DESCRIPTIONS.get(
                class_name, f"Checklist class from data_type: {class_name}."
            ),
            "slots": sorted(class_slots[class_name]),
        }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as handle:
        yaml.dump(classes_doc, handle, sort_keys=False, allow_unicode=True)

    print(f"Wrote {len(classes_doc['classes'])} classes to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
