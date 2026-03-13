
import yaml
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from collections import defaultdict
from column_order import column_order

# Load schema
with open("schema.yaml", "r") as f:
    schema = yaml.safe_load(f)

slots = schema.get("slots", {})
classes = schema.get("classes", {})

slot_to_classes = defaultdict(list)
if isinstance(classes, dict):
    for class_name, class_def in classes.items():
        if not isinstance(class_def, dict):
            continue
        for slot_name in class_def.get("slots", []) or []:
            slot_to_classes[slot_name].append(class_name)


def slot_sections(slot):
    subsets = slot.get("in_subset", [])
    if isinstance(subsets, list):
        return [s for s in subsets if s]
    if isinstance(subsets, str) and subsets.strip():
        return [subsets.strip()]
    legacy = slot.get("annotations", {}).get("section", "")
    return [legacy] if legacy else []

# Extract version from schema
schema_version = schema.get("version", "unknown")

# Setup workbook
wb = Workbook()
ws_readme = wb.active
ws_readme.title = "README"

# README content
ws_readme["A1"] = f"These templates were generated using the FAIR eDNA checklist version {schema_version}"
ws_readme["A4"] = "Date/Time generated;"
ws_readme["A5"] = datetime.now().isoformat()
ws_readme["A7"] = (
    "Note: This template includes all data components and terms from the FAIR eDNA checklist, "
    "encompassing sections for both targeted assays and metabarcoding assay types.")
ws_readme["A9"] = "Requirement levels;"
readme_levels = [
    ("A10", "M = Mandatory", "E26B0A"),
    ("A11", "HR = Highly recommended", "F6CD46"),
    ("A12", "R = Recommended", "FFFFA6"),
    ("A13", "O = Optional", "D6FDA4")
]

def make_fill(hex_color):
    return PatternFill(start_color="FF" + hex_color, end_color="FF" + hex_color, fill_type="solid")

for cell, text, color in readme_levels:
    ws_readme[cell] = text
    ws_readme[cell].fill = make_fill(color)

level_colors = {
    "M": "E26B0A",
    "HR": "F6CD46",
    "R": "FFFFA6",
    "O": "D6FDA4"
}
bold_font = Font(bold=True)

# ProjectMetadata tab
ws_proj = wb.create_sheet("projectMetadata")
ws_proj.append(["requirement_level_code", "section", "term_name", "project_level"])
for col in range(1, 5):
    ws_proj.cell(row=1, column=col).font = bold_font

for slot_name, slot in slots.items():
    slot_classes = [c for c in slot_to_classes.get(slot_name, []) if c != "MetadataChecklist"]
    if "projectMetadata" not in slot_classes:
        continue
    annotations = slot.get("annotations", {})
    req = annotations.get("requirement_level_code", "")
    section = " | ".join(slot_sections(slot))
    ws_proj.append([req, section, slot_name, ""])
    row = ws_proj.max_row
    if req in level_colors:
        ws_proj.cell(row=row, column=1).fill = make_fill(level_colors[req])
    ws_proj.cell(row=row, column=3).font = bold_font
    def first_example_value(s):
        ex = s.get("examples", []) or []
        if isinstance(ex, list) and ex:
            item = ex[0]
            if isinstance(item, dict):
                return item.get("value", "")
            return str(item)
        return ""
    comment_text = "\n".join(filter(None, [
        slot.get("description", ""),
        f"Example: {first_example_value(slot)}" if first_example_value(slot) else "",
        f"Requirement: {req}" if req else "",
        f"Field type: {slot.get('range')}" if slot.get("range") else ""
    ]))
    if comment_text:
        ws_proj.cell(row=row, column=3).comment = Comment(comment_text, "FAIRe")

# Prefix values for column A3 in selected tabs
targeted_prefix = {
    "sampleMetadata": "samp_name",
    "ampData": "samp_name",
    "stdData": "samp_name",
    "experimentRunMetadata": "samp_name",
    "eLowQuantData": "assay_name",
    "taxaRaw": "seq_id",
    "taxaFinal": "seq_id"
}

# Other data_type tabs
data_type_slots = defaultdict(list)
for class_name, class_def in classes.items():
    if class_name == "MetadataChecklist" or not isinstance(class_def, dict):
        continue
    for slot_name in class_def.get("slots", []) or []:
        slot = slots.get(slot_name)
        if slot is not None:
            data_type_slots[class_name].append((slot_name, slot))

for dt, slot_list in sorted(data_type_slots.items()):
    if dt == "projectMetadata":
        continue

    prefix = targeted_prefix.get(dt, "")

    if dt in column_order:
        slot_dict = {name: slot for name, slot in slot_list}
        sorted_slot_list = [(name, slot_dict[name]) for name in column_order[dt] if name in slot_dict]
    else:
        sorted_slot_list = sorted(slot_list, key=lambda x: x[0])

    # Remove prefix column if it would be duplicated
    if prefix and prefix in [name for name, _ in sorted_slot_list]:
        sorted_slot_list = [(n, s) for n, s in sorted_slot_list if n != prefix]

    ws = wb.create_sheet(title=dt)
    ws.append(["# requirement_level_code"] + [slot.get("annotations", {}).get("requirement_level_code", "") for _, slot in sorted_slot_list])
    ws.append(["# section"] + [" | ".join(slot_sections(slot)) for _, slot in sorted_slot_list])
    ws.append([prefix] + [name for name, _ in sorted_slot_list])
    for col, (slot_name, slot) in enumerate(sorted_slot_list, start=1):
        col_letter = get_column_letter(col + 1)  # shift by 1 due to prefix column
        level = slot.get("annotations", {}).get("requirement_level_code", "")
        if level in level_colors:
            for row in range(1, 4):
                ws[f"{col_letter}{row}"].fill = make_fill(level_colors[level])
        ws[f"{col_letter}3"].font = bold_font
        comment_text = "\n".join(filter(None, [
            slot.get("description", ""),
            f"Example: {first_example_value(slot)}" if first_example_value(slot) else "",
            f"Requirement: {level}" if level else ""
        ]))
        if comment_text:
            ws[f"{col_letter}3"].comment = Comment(comment_text, "FAIRe")

# Drop-down values tab with expansion
dropdown_rows = []
max_options = 0

for slot_name, slot in slots.items():
    enum_values = slot.get("enum_values", {})
    is_cv = isinstance(enum_values, dict) and bool(enum_values)
    is_boolean = slot.get("range") == "boolean"
    if not is_cv and not is_boolean:
        continue
    if isinstance(enum_values, dict):
        options = list(enum_values.keys())
    elif isinstance(enum_values, list):
        options = [str(v) for v in enum_values]
    elif is_boolean:
        options = ["true", "false"]
    else:
        options = []
    term_type = "controlled vocabulary" if is_cv else "Boolean"
    max_options = max(max_options, len(options))
    row = [slot_name, term_type, len(options), "|".join(options)] + options
    dropdown_rows.append(row)

ws_dd = wb.create_sheet("Drop-down values")
base_headers = ["term_name", "term_type", "n_options", "vocab_options"]
vocab_headers = [f"vocab{i+1}" for i in range(max_options)]
headers = base_headers + vocab_headers
ws_dd.append(headers)
for col in range(1, len(headers) + 1):
    ws_dd.cell(row=1, column=col).font = bold_font

for row in dropdown_rows:
    ws_dd.append(row)

# Optional: reorder the tabs to a desired sequence
desired_order = [
    "README",
    "projectMetadata",
    "Drop-down values",
    "sampleMetadata",
    "ampData",
    "stdData",
    "experimentRunMetadata",
    "eLowQuantData",
    "taxaRaw",
    "taxaFinal"
]

# Reorder sheets (only works with openpyxl >= 2.5)
wb._sheets.sort(key=lambda ws: desired_order.index(ws.title) if ws.title in desired_order else 999)

# Save final Excel file
output_filename = f"FAIRe_checklist_v{schema_version}_FULLtemplate.xlsx"
wb.save(output_filename)

