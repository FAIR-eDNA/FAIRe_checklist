import os

import yaml
from openpyxl import load_workbook
from checklist_source import resolve_single_checklist_file

# Builds enums from the FAIRe checklist
SLOTS_DIR = "slots"
ENUMS_PATH = "enums.yaml"


def split_pipe(raw):
    if not raw:
        return []
    return [x.strip() for x in str(raw).split("|") if x.strip()]


def load_checklist_rows_from_xlsx(path, sheet_name="checklist"):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Checklist Excel file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    header_names = [str(h).strip() if h is not None else "" for h in headers]
    out = []
    for row in rows:
        rec = {}
        for idx, key in enumerate(header_names):
            rec[key] = row[idx] if idx < len(row) else None
        out.append(rec)
    return out


def slot_files():
    return sorted(
        f
        for f in os.listdir(SLOTS_DIR)
        if f.endswith(".yaml") and f not in {"glossary_annotation.yaml"}
    )


def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def ensure_enum(enums, enum_name, slot_name):
    if enum_name not in enums:
        enums[enum_name] = {
            "description": f"Controlled vocabulary for {slot_name}.",
            "permissible_values": {},
        }
    if "permissible_values" not in enums[enum_name] or not isinstance(
        enums[enum_name]["permissible_values"], dict
    ):
        enums[enum_name]["permissible_values"] = {}
    return enums[enum_name]["permissible_values"]


def add_pv(pv, value, meaning=None):
    if not value:
        return
    if value not in pv:
        pv[value] = {"meaning": meaning or value}


def main():
    checklist_xlsx = resolve_single_checklist_file(".xlsx")
    enums = {}
    enums["MissingValueEnum"] = {
        "description": "A controlled vocabulary for reporting why a value is missing.",
        "annotations": {
            "source": "https://terminology.hl7.org/3.0.0/ValueSet-v3-NullFlavor.html"
        },
        "permissible_values": {"unknown": {"description": "TODO: map preferred null-flavor codes."}},
    }

    # Primary source: checklist Excel controlled vocabulary options.
    for row in load_checklist_rows_from_xlsx(checklist_xlsx):
        slot_name = (row.get("term_name") or "").strip()
        if not slot_name:
            continue
        cv = split_pipe(row.get("controlled_vocabulary_options"))
        if not cv:
            continue
        enum_name = f"{slot_name}_enum"
        pv = ensure_enum(enums, enum_name, slot_name)
        for v in cv:
            add_pv(pv, v)

    # Merge in any extra existing values from slot-local enum_values so we don't drop old content.
    for file_name in slot_files():
        slot = load_yaml(os.path.join(SLOTS_DIR, file_name))
        if not isinstance(slot, dict) or "name" not in slot:
            continue
        ev = slot.get("enum_values")
        if not isinstance(ev, dict) or not ev:
            continue
        enum_name = slot.get("range")
        if not isinstance(enum_name, str) or not enum_name.endswith("_enum"):
            enum_name = f"{slot['name']}_enum"
        pv = ensure_enum(enums, enum_name, slot["name"])
        for key, val in ev.items():
            if isinstance(val, dict):
                add_pv(pv, key, val.get("meaning", key))
            else:
                add_pv(pv, key, key)

    sorted_enums = {}
    for enum_name in sorted(enums.keys()):
        enum_def = enums[enum_name]
        pv = enum_def.get("permissible_values", {})
        sorted_pv = {}
        for pv_key in sorted(pv.keys()):
            sorted_pv[pv_key] = pv[pv_key]
        enum_def["permissible_values"] = sorted_pv
        sorted_enums[enum_name] = enum_def

    final_doc = {"enums": sorted_enums}
    with open(ENUMS_PATH, "w", encoding="utf-8") as f:
        yaml.dump(final_doc, f, sort_keys=False, allow_unicode=True)

    print(f"Wrote {len(final_doc['enums'])} enums to {ENUMS_PATH}")


if __name__ == "__main__":
    main()
