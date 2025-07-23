import yaml
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# === Load glossary_annotation.yaml ===
with open("slots/glossary_annotation.yaml", "r") as f:
    glossary_data = yaml.safe_load(f)

# === Extract glossary block safely from nested structure ===
if (
    isinstance(glossary_data, dict)
    and "annotations" in glossary_data
    and isinstance(glossary_data["annotations"], dict)
    and "glossary" in glossary_data["annotations"]
    and isinstance(glossary_data["annotations"]["glossary"], dict)
):
    # Correct glossary extraction
    glossary_block = glossary_data.get("annotations", {}).get("glossary", {})
    if not isinstance(glossary_block, dict):
        raise ValueError("Expected annotations.glossary to be a dictionary of term: definition pairs")
    glossary = glossary_block

else:
    raise ValueError("Expected 'annotations.glossary' block in the YAML.")

# === Load schema.yaml ===
with open("schema.yaml", "r") as f:
    schema = yaml.safe_load(f)
slots = schema.get("slots", {})

# === Ordered columns ===
ordered_columns = [
    "data_type", "section", "term_name", "description", "requirement_level_code",
    "requirement_level", "requirement_level_condition", "term_type", "unit", "fixed_format",
    "controlled_vocabulary_options", "example", "sample_type_specificity", "source",
    "URI", "modifications_made"
]

# === Custom section order and colors ===
custom_section_order = [
    "Project", "Data management", "Sample information", "Sample relations", "Sample collection",
    "Sample storage", "Sample preparation", "Nucleic acid extraction", "PCR",
    "Targeted assay detection", "Library preparation sequencing", "Bioinformatics",
    "OTU/ASV", "Environment"
]

requirement_specific_colors = {
    "M": "FFE26B0A", "HR": "FFFFCC00", "R": "FFFDFD96", "O": "FFCCFF99"
}

explicit_section_colors = {
    "Project": "FFFBB4AE", "Data management": "FFB3CDE3", "Sample information": "FFCCEBC5",
    "Sample relations": "FFDECBE4", "Sample collection": "FFFED9A6", "Sample storage": "FFFFFFCC",
    "Sample preparation": "FFE5D8BD", "Nucleic acid extraction": "FFFDDAEC", "PCR": "FFFBB4AE",
    "Targeted assay detection": "FFB3CDE3", "Library preparation sequencing": "FFCCEBC5",
    "Bioinformatics": "FFDECBE4", "OTU/ASV": "FFFED9A6", "Environment": "FFFFFF99"
}

# === Build checklist DataFrame ===
records = []
for slot_name, slot in slots.items():
    annotations = slot.get("annotations", {})
    def get_annot(field): return " | ".join(annotations[field]) if isinstance(annotations.get(field), list) else annotations.get(field, "")
    enum_values = slot.get("enum_values", {})
    is_enum = isinstance(enum_values, dict) and enum_values
    term_type = "controlled vocabulary" if is_enum else slot.get("range", "")
    enum_opts = " | ".join(enum_values.keys()) if is_enum else ""
    record = {
        "data_type": get_annot("data_type"),
        "section": get_annot("section"),
        "term_name": slot_name,
        "description": slot.get("description", ""),
        "requirement_level_code": get_annot("requirement_level_code"),
        "requirement_level": get_annot("requirement_level"),
        "requirement_level_condition": get_annot("requirement_level_condition"),
        "term_type": term_type,
        "unit": get_annot("unit"),
        "fixed_format": get_annot("fixed_format"),
        "controlled_vocabulary_options": enum_opts,
        "example": get_annot("example"),
        "sample_type_specificity": get_annot("sample_type_specificity"),
        "source": get_annot("source"),
        "URI": slot.get("slot_uri", ""),
        "modifications_made": get_annot("modifications_made"),
    }
    records.append(record)

checklist_df = pd.DataFrame(records)[ordered_columns]
checklist_df["section"] = pd.Categorical(checklist_df["section"], categories=custom_section_order, ordered=True)
checklist_df = checklist_df.sort_values(by=["section", "data_type", "term_name"], na_position="last")

# === Create Excel workbook ===
wb = Workbook()
ws_readme = wb.active
ws_readme.title = "README"
bold = Font(bold=True)

# === Write instructions ===
instructions = [
    'Date and time must be recorded following ISO 8601 format (yyyy-mm-ddThh:mm:ss)...',
    "Time duration must be recorded following ISO 8601 durations (PnYnMnWnDTnHnMnS...)",
    "A date and time range can be specified using a forward slash (/) ...",
    "Use space vertical bar space ( | ) to separate multiple values in a list...",
    "Use hyphen (-) to indicate a range of numeric or integer values...",
    "For numeric fields, enter only numbers and do not enter units...",
    "For Boolean type of questions, always read the descriptions and answer with 0 or 1.",
    'When "other" is selected..., write the answer using the format of "other: FREE TEXT DESCRIPTION".',
    "If suitable term names are not available..., users should search for them in MIxS or DwC...",
    "When a value is missing from a mandatory term..., use the INSDC missing value vocabulary..."
]

ws_readme.cell(row=1, column=1, value="Instructions for data submitters:").font = bold
for i, line in enumerate(instructions, start=2):
    ws_readme.cell(row=i, column=2, value=line)

# === Write glossary ===
start_row = len(instructions) + 3
ws_readme.cell(row=start_row, column=1, value="Terms and definitions:").font = bold
ws_readme.cell(row=start_row + 1, column=2, value="Term").font = bold
ws_readme.cell(row=start_row + 1, column=3, value="Definition").font = bold
for i, (term, definition) in enumerate(glossary.items(), start=start_row + 2):
    ws_readme.cell(row=i, column=2, value=str(term))
    ws_readme.cell(row=i, column=3, value=str(definition))

# === Add checklist tab ===
ws_checklists = wb.create_sheet("checklist")
for col_num, col_name in enumerate(ordered_columns, start=1):
    cell = ws_checklists.cell(row=1, column=col_num, value=col_name)
    cell.font = bold

for row_idx, row in enumerate(checklist_df.itertuples(index=False), start=2):
    row_vals = list(row)
    section_val = row_vals[1]
    req_code_val = row_vals[4]
    section_fill = PatternFill(start_color=explicit_section_colors.get(section_val, ""),
                               end_color=explicit_section_colors.get(section_val, ""),
                               fill_type="solid") if section_val else None
    requirement_fill = PatternFill(start_color=requirement_specific_colors.get(req_code_val, ""),
                                   end_color=requirement_specific_colors.get(req_code_val, ""),
                                   fill_type="solid") if req_code_val else None
    for col_idx, value in enumerate(row_vals, start=1):
        cell = ws_checklists.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
        if col_idx == 2 and section_fill:
            cell.fill = section_fill
        elif col_idx in (5, 6) and requirement_fill:
            cell.fill = requirement_fill

# === Save final workbook ===
output_path = "FAIRe_checklist_v1.0.2_test.xlsx"
wb.save(output_path)
print(f"✅ Excel file written: {output_path}")
