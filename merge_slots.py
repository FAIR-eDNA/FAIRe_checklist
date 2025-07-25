import yaml
from openpyxl import Workbook
from openpyxl.styles import Font

# === Load and validate glossary YAML ===
with open("slots/glossary_annotation.yaml", "r") as f:
    data = yaml.safe_load(f)

# Handle case where top-level key is 'glossary', or it's already flat
if "glossary" in data:
    glossary = data["glossary"]
else:
    glossary = data

if not isinstance(glossary, dict):
    raise ValueError("Expected glossary to be a dictionary of term: definition pairs.")

# === Setup workbook ===
wb = Workbook()
ws = wb.active
ws.title = "README"
bold = Font(bold=True)

# === Add instructions section ===
instructions = [
    'Date and time must be recorded following ISO 8601 format (yyyy-mm-ddThh:mm:ss). Time zone must be specified after the timestamp e.g., "2008-01-23T19:23-06:00" in the time zone six hours earlier than UTC, or "2008-01-23T19:23Z" at UTC time. In Excel, format the cell as text to prevent from automatic conversion to other date fromats.',
    "Time duration must be recorded following ISO 8601 durations (PnYnMnWnDTnHnMnS for P<date>T<time>). e.g., P1Y1M1DT1H1M1.1S = One year, one month, one day, one hour, one minute, one second, and 100 milliseconds",
    "A date and time range can be specified using a forward slash (/) to separate the start and end values (e.g., 2008-01-23T19:23-06:00/2008-01-23T19:53-06:00).",
    "Use space vertical bar space ( | ) to separate multiple values in a list. i.e., x | y | z",
    "Use hyphen (-) to indicate a range of numeric or integer values. In Excel, format the cell as text to prevent it from being auto-formatted as a date.",
    "For numeric fields, enter only numbers and do not enter units. Read the descriptions for the standard unit.",
    "For Boolean type of questions, always read the descriptions and answer with 0 or 1.",
    'When "other" is selected in a controlled vocabulary term, write the answer using the format of "other: FREE TEXT DESCRIPTION".',
    "If suitable term names are not available in the current checklist, users should search for them in existing standards, such as MIxS (https://genomicsstandardsconsortium.github.io/mixs/) and DwC (https://dwc.tdwg.org/terms/), and use these standardized terms where possible. If relevant terms cannot be found in these resources, users may add new terms using clear, concise, and descriptive names within related tables.",
    "When a value is missing from a mandatory term, it is required to provide the reason using the INSDC missing value controlled vocabulary (https://www.insdc.org/submitting-standards/missing-value-reporting/). See https://fair-edna.github.io/guidelines.html#missing-values for the full list of values."
]

# Section header
ws.cell(row=1, column=1, value="Instructions for data submitters:").font = bold
for i, line in enumerate(instructions, start=2):
    ws.cell(row=i, column=2, value=line)

# === Add glossary section ===
glossary_start = len(instructions) + 3
ws.cell(row=glossary_start, column=1, value="Terms and definitions:").font = bold
ws.cell(row=glossary_start + 1, column=2, value="Term").font = bold
ws.cell(row=glossary_start + 1, column=3, value="Definition").font = bold

for i, (term, definition) in enumerate(glossary.items(), start=glossary_start + 2):
    ws.cell(row=i, column=2, value=str(term))
    ws.cell(row=i, column=3, value=str(definition))

# Save file
wb.save("faire_readme_formatted.xlsx")
print("✅ Excel file written: faire_readme_formatted.xlsx")
